#!/usr/bin/python3

import argparse
import pymongo
from openpyxl import Workbook
import threading


def print_error(message):
    print("[!] %s" % message)


def print_info(message):
    print("[ ] %s" % message)


def get_collection_names():
    return db.list_collection_names()


def get_data_from_collection(collection_name):
    return db[collection_name].find({}, {"_id": False})


def save_data_to_workbook(sheet, data):
    sheet.append(list(dict.keys(data[0])))

    for record in data:
        sheet.append(list(dict.values(record)))


parser = argparse.ArgumentParser(
    description="Save MongoDB database collections data to xlsx")

parser.add_argument("--host", required=False,
                    default="127.0.0.1", help="MongoDB host")
parser.add_argument("--port", required=False,
                    default=27017, help="MongoDB port")
parser.add_argument("--db", required=True, help="MongoDB database to dump")
parser.add_argument("--out", required=True, help="output file path")

args = parser.parse_args()

HOST = args.host
PORT = args.port
DB = args.db
OUT = args.out
MAX_WORKERS = 5

print_info("Connecting to MongoDB")
client = pymongo.MongoClient(HOST, PORT)
db = client[DB]
wb = Workbook()

collection_names = get_collection_names()
for name in collection_names:
    print_info("Saving collection %s" % name)
    iterator = get_data_from_collection(collection_name=name)
    sheet = wb.create_sheet(title=name)
    save_data_to_workbook(sheet=sheet, data=iterator)
    print_info("DONE")

# if 'Sheet' in wb.sheetnames:
#    wb.remove_sheet('Sheet')

print_info("Saving spreadsheet")
wb.save(filename=OUT)
